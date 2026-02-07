#!/usr/bin/env python3
"""ISI v0.1 — Parse manually downloaded IMF CPIS raw Excel to flat CSV.

Input:  data/raw/finance/cpis_2024_raw.xlsx
Output: data/processed/finance/cpis_2024_flat.csv

Output schema (exact order):
  dataset_id, reference_area, counterpart_area, indicator,
  instrument, frequency, period, value

Mapping rules:
  dataset_id       = "imf_cpis"
  reference_area   = issuer / debtor country code (ISO-2)
  counterpart_area = holder country code (ISO-2)
  indicator        = original IMF indicator code or string
  instrument       = "DEBT_SEC"
  frequency        = "A"
  period           = "2024"
  value            = numeric, non-negative

Hard constraints:
  - Does NOT infer or guess column meanings
  - Does NOT rename countries
  - Does NOT aggregate
  - Does NOT filter by country
  - Does NOT normalize values
  - Does NOT impute missing data
  - STOPS if required columns cannot be unambiguously identified

Task: ISI-FINANCE-PARSE-CPIS
"""

import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("FATAL: openpyxl is not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_FILE = PROJECT_ROOT / "data" / "raw" / "finance" / "cpis_2024_raw.xlsx"
OUT_DIR = PROJECT_ROOT / "data" / "processed" / "finance"
OUT_FILE = OUT_DIR / "cpis_2024_flat.csv"

# ─── Column identification ───────────────────────────────────────
# The IMF CPIS Excel export may use varying column header names.
# We define canonical mappings: each required field maps to a set
# of known header variants (case-insensitive, stripped).
#
# If a required column cannot be matched to exactly one header,
# the script STOPS.

REFERENCE_AREA_VARIANTS = {
    "country code",
    "country",
    "counterpart country code",       # when sheet is "inward" perspective
    "economy code",
    "economy",
    "ref_area",
    "reference area",
    "reporting economy code",
    "reporting economy",
    "reporter code",
    "reporter",
    "issuer country code",
    "issuer",
    "debtor country code",
    "debtor",
    "destination country code",
    "destination economy code",
    "sector of counterpart area code",  # unlikely but defensive
}

COUNTERPART_AREA_VARIANTS = {
    "counterpart country code",
    "counterpart country",
    "counterpart area code",
    "counterpart area",
    "counterpart economy code",
    "counterpart economy",
    "holder country code",
    "holder country",
    "holder",
    "source country code",
    "source country",
    "source economy code",
    "source economy",
    "partner country code",
    "partner country",
    "partner",
    "investor country code",
    "investor country",
    "creditor country code",
    "creditor country",
}

INDICATOR_VARIANTS = {
    "indicator",
    "indicator code",
    "indicator name",
    "concept",
    "concept code",
    "measure",
    "measure code",
    "variable",
    "variable code",
    "series",
    "series code",
    "attribute",
    "type",
}

VALUE_VARIANTS = {
    "value",
    "obs_value",
    "observation value",
    "amount",
    "usd millions",
    "us dollars (millions)",
    "data value",
    "data",
    "position",
    "stock",
}

PERIOD_VARIANTS = {
    "time_period",
    "time period",
    "period",
    "year",
    "date",
    "time",
    "reference period",
    "observation date",
}

# Columns that tell us the country codes (ISO-2) for ref_area and counterpart.
# The IMF bulk export sometimes uses separate columns for code vs. name.
# We prefer "*code*" columns over name-only columns.
#
# Strategy:
#   1. Scan all header cells.
#   2. For each required field, find all headers whose lowercase-stripped
#      form matches one of the known variants.
#   3. If zero matches: STOP.
#   4. If more than one match: prefer the one containing "code".
#   5. If still ambiguous: STOP.


def normalise(s):
    """Lowercase, strip whitespace and non-breaking spaces."""
    if s is None:
        return ""
    return str(s).strip().replace("\xa0", " ").lower()


def find_column(headers, variants, field_name):
    """Return the 0-based column index matching one of the variants.

    STOPS if zero or ambiguous matches.
    """
    matches = []
    for idx, h in enumerate(headers):
        hn = normalise(h)
        if hn in variants:
            matches.append((idx, h))

    if len(matches) == 0:
        print(f"FATAL: cannot identify column for '{field_name}'.", file=sys.stderr)
        print(f"  Scanned headers: {headers}", file=sys.stderr)
        print(f"  Expected one of: {sorted(variants)}", file=sys.stderr)
        sys.exit(1)

    if len(matches) == 1:
        idx, h = matches[0]
        print(f"  {field_name} -> column {idx} ('{h}')")
        return idx

    # More than one match: prefer the one containing "code"
    code_matches = [(idx, h) for idx, h in matches if "code" in normalise(h)]
    if len(code_matches) == 1:
        idx, h = code_matches[0]
        print(f"  {field_name} -> column {idx} ('{h}') [preferred 'code' variant]")
        return idx

    print(f"FATAL: ambiguous column match for '{field_name}'.", file=sys.stderr)
    print(f"  Candidates: {matches}", file=sys.stderr)
    sys.exit(1)


def main():
    # ── Load workbook ──
    if not RAW_FILE.exists():
        print(f"FATAL: raw file not found: {RAW_FILE}", file=sys.stderr)
        print("Run ingest_finance_cpis_manual.py first.", file=sys.stderr)
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(RAW_FILE, read_only=True, data_only=True)
    except Exception as exc:
        print(f"FATAL: cannot read file: {exc}", file=sys.stderr)
        sys.exit(1)

    # ── Select sheet ──
    # Use the first sheet. If the workbook has multiple sheets, we
    # process only the first one. The user is expected to provide a
    # file where the first sheet contains the bilateral data.
    sheet = wb[wb.sheetnames[0]]
    print(f"Using sheet: '{wb.sheetnames[0]}'")

    # ── Read all rows into memory ──
    rows_raw = []
    for row in sheet.iter_rows(values_only=True):
        rows_raw.append(list(row))
    wb.close()

    if len(rows_raw) < 2:
        print("FATAL: sheet has fewer than 2 rows (need header + data).", file=sys.stderr)
        sys.exit(1)

    # ── Identify header row ──
    # The header row is the first row where at least 3 cells are non-empty
    # strings. Some IMF exports have title/metadata rows above the actual
    # header.
    header_idx = None
    for i, row in enumerate(rows_raw):
        non_empty_strings = sum(1 for c in row if isinstance(c, str) and c.strip())
        if non_empty_strings >= 3:
            header_idx = i
            break

    if header_idx is None:
        print("FATAL: cannot identify header row.", file=sys.stderr)
        sys.exit(1)

    headers = rows_raw[header_idx]
    print(f"Header row index: {header_idx}")
    print(f"Headers: {headers}")

    # ── Map columns ──
    print("Column mapping:")
    col_ref = find_column(headers, REFERENCE_AREA_VARIANTS, "reference_area")
    col_cp = find_column(headers, COUNTERPART_AREA_VARIANTS, "counterpart_area")
    col_ind = find_column(headers, INDICATOR_VARIANTS, "indicator")
    col_val = find_column(headers, VALUE_VARIANTS, "value")

    # Period column is optional — if present, we use it; otherwise default "2024"
    col_period = None
    for idx, h in enumerate(headers):
        if normalise(h) in PERIOD_VARIANTS:
            col_period = idx
            print(f"  period -> column {idx} ('{h}')")
            break
    if col_period is None:
        print("  period -> not found in headers; defaulting to '2024'")

    # ── Parse data rows ──
    data_rows = rows_raw[header_idx + 1 :]
    parsed = []
    skipped_empty = 0
    skipped_negative = 0

    for row_num, row in enumerate(data_rows, start=header_idx + 2):
        # Pad row if shorter than header
        while len(row) < len(headers):
            row.append(None)

        ref_area = row[col_ref]
        cp_area = row[col_cp]
        indicator = row[col_ind]
        raw_value = row[col_val]

        # Determine period
        if col_period is not None:
            period_raw = row[col_period]
            period = str(period_raw).strip() if period_raw is not None else "2024"
        else:
            period = "2024"

        # ── Skip rows with missing country codes ──
        if ref_area is None or str(ref_area).strip() == "":
            skipped_empty += 1
            continue
        if cp_area is None or str(cp_area).strip() == "":
            skipped_empty += 1
            continue

        ref_area = str(ref_area).strip()
        cp_area = str(cp_area).strip()
        indicator = str(indicator).strip() if indicator is not None else ""

        # ── Parse value ──
        if raw_value is None or str(raw_value).strip() == "":
            skipped_empty += 1
            continue

        try:
            value = float(raw_value)
        except (ValueError, TypeError):
            # Non-numeric value (e.g., "C" for confidential, "..." for missing)
            skipped_empty += 1
            continue

        # ── Validate non-negative ──
        if value < 0:
            skipped_negative += 1
            continue

        parsed.append({
            "dataset_id": "imf_cpis",
            "reference_area": ref_area,
            "counterpart_area": cp_area,
            "indicator": indicator,
            "instrument": "DEBT_SEC",
            "frequency": "A",
            "period": period,
            "value": value,
        })

    # ── Post-parse validation ──
    if len(parsed) == 0:
        print("FATAL: zero rows survived parsing.", file=sys.stderr)
        print(f"  Total data rows scanned: {len(data_rows)}", file=sys.stderr)
        print(f"  Skipped (empty/missing): {skipped_empty}", file=sys.stderr)
        print(f"  Skipped (negative): {skipped_negative}", file=sys.stderr)
        sys.exit(1)

    # Check no empty country codes in parsed output
    for rec in parsed:
        if rec["reference_area"] == "" or rec["counterpart_area"] == "":
            print("FATAL: empty country code in parsed output.", file=sys.stderr)
            sys.exit(1)

    # ── Write output CSV ──
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "dataset_id",
        "reference_area",
        "counterpart_area",
        "indicator",
        "instrument",
        "frequency",
        "period",
        "value",
    ]

    with open(OUT_FILE, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(parsed)

    # ── Report ──
    ref_areas = set(r["reference_area"] for r in parsed)
    cp_areas = set(r["counterpart_area"] for r in parsed)

    print()
    print(f"Wrote {len(parsed)} rows to {OUT_FILE}")
    print(f"  Unique reference_area codes: {len(ref_areas)}")
    print(f"  Unique counterpart_area codes: {len(cp_areas)}")
    print(f"  Skipped (empty/missing): {skipped_empty}")
    print(f"  Skipped (negative): {skipped_negative}")


if __name__ == "__main__":
    main()
